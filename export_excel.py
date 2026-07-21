"""Excel export — data/blinkit_review_analysis.xlsx (pandas + openpyxl).

Three tabs, screenshot-friendly (frozen header rows, sensible column widths, bold
section titles) so tables can go straight into a deck:
  1. Tagged Reviews — one row per tagged item, all fields as columns.
  2. Insights      — the Pass-2 synthesis (barriers %, drivers, signal
                     distribution, segments, JTBDs, unmet needs, quotes,
                     hypotheses w/ evidence + confidence).
  3. Validation    — coverage/ambiguity, per-theme evidence + triangulation,
                     per-hypothesis validation, single-source flags.

Regenerated on every full pipeline run.

Usage:  python export_excel.py
Output: data/blinkit_review_analysis.xlsx
"""

import json
import os

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from common import DATA_DIR, get_logger, read_jsonl

log = get_logger("excel")

TAGGED_PATH = os.path.join(DATA_DIR, "tagged.jsonl")
SYNTH_PATH = os.path.join(DATA_DIR, "synthesis.json")
VALIDATION_PATH = os.path.join(DATA_DIR, "validation.json")
XLSX_PATH = os.path.join(DATA_DIR, "blinkit_review_analysis.xlsx")

TITLE_FONT = Font(bold=True, size=13, color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
WRAP = Alignment(wrap_text=True, vertical="top")


def _join(v):
    return ", ".join(v) if isinstance(v, list) else v


# ── Tab 1: Tagged Reviews ───────────────────────────────────────────

def tagged_df(tagged: list[dict]) -> pd.DataFrame:
    cols = ["id", "source", "date", "rating", "author", "is_relevant", "sentiment",
            "exploration_signal", "confidence", "themes", "user_segment_signals",
            "mentions_category", "job_to_be_done", "frustration_root_cause",
            "direct_quote", "text", "url"]
    rows = []
    for it in tagged:
        row = {c: it.get(c) for c in cols}
        for lst in ("themes", "user_segment_signals", "mentions_category"):
            row[lst] = _join(it.get(lst) or [])
        rows.append(row)
    return pd.DataFrame(rows, columns=cols)


# ── Sheet-building helpers (openpyxl) ───────────────────────────────

def _write_title(ws, row, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font = TITLE_FONT
    return row + 1


def _write_table(ws, row, df: pd.DataFrame, wrap_cols=()):
    """Write a DataFrame as a styled table starting at `row`. Returns next free row."""
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=row, column=j, value=str(col))
        c.font, c.fill = HEADER_FONT, HEADER_FILL
    for i, (_, r) in enumerate(df.iterrows(), start=row + 1):
        for j, col in enumerate(df.columns, start=1):
            val = r[col]
            if isinstance(val, list):
                val = _join(val)
            c = ws.cell(row=i, column=j, value=val)
            if col in wrap_cols:
                c.alignment = WRAP
    return row + len(df) + 2  # blank line after


def _autosize(ws, widths: dict):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


# ── Tab 2: Insights ─────────────────────────────────────────────────

def build_insights(ws, s: dict):
    a = s["aggregates"]
    c = a["counts"]
    row = _write_title(ws, 1, "Blinkit Category-Exploration — Insights")
    ws.cell(row=row, column=1,
            value=f"Generated {s['generated_at'][:16]} · {c['relevant']} relevant of "
                  f"{c['total_tagged']} tagged ({c['relevant_pct']}%) · "
                  f"{c['irrelevant_filtered']} filtered as noise")
    row += 1
    ws.cell(row=row, column=1, value=s["framing"]).font = Font(italic=True)
    row += 2

    row = _write_title(ws, row, "Top barriers to category exploration")
    row = _write_table(ws, row, pd.DataFrame(a["barriers"][:8]))

    row = _write_title(ws, row, "What's working (drivers)")
    row = _write_table(ws, row, pd.DataFrame(a.get("drivers", [])))

    row = _write_title(ws, row, "Exploration-signal distribution")
    row = _write_table(ws, row, pd.DataFrame(a["exploration_signal_distribution"]))

    row = _write_title(ws, row, "Segments most locked into repetitive buying")
    row = _write_table(ws, row, pd.DataFrame(s["locked_in_segments"]), wrap_cols=("traits",))

    row = _write_title(ws, row, "Segments most likely to explore")
    row = _write_table(ws, row, pd.DataFrame(s["explorer_segments"]), wrap_cols=("traits",))

    row = _write_title(ws, row, "Top jobs-to-be-done")
    row = _write_table(ws, row, pd.DataFrame(s["top_jobs_to_be_done"]), wrap_cols=("job",))

    row = _write_title(ws, row, "Top unmet needs")
    row = _write_table(ws, row, pd.DataFrame({"unmet_need": s["unmet_needs"]}), wrap_cols=("unmet_need",))

    row = _write_title(ws, row, "Most powerful quotes")
    row = _write_table(ws, row, pd.DataFrame(s["powerful_quotes"]), wrap_cols=("quote", "attribution"))

    row = _write_title(ws, row, "Hypotheses (with evidence strength)")
    hyp = [{"hypothesis": h["hypothesis"], "confidence": h["confidence"],
            "evidence_items": h["evidence_items"],
            "n_sources": len(h.get("evidence_sources", [])),
            "sources": _join(h.get("evidence_sources", [])),
            "triangulated": h["source_triangulated"],
            "rationale": h.get("rationale", "")} for h in s["hypotheses"]]
    _write_table(ws, row, pd.DataFrame(hyp), wrap_cols=("hypothesis", "rationale"))

    _autosize(ws, {"A": 46, "B": 16, "C": 14, "D": 12, "E": 22, "F": 14, "G": 60})
    ws.freeze_panes = "A2"


# ── Tab 3: Validation ───────────────────────────────────────────────

def build_validation(ws, v: dict):
    row = _write_title(ws, 1, "Insight-Quality Validation")
    ws.cell(row=row, column=1, value=v["purpose"]).font = Font(italic=True)
    row += 2

    cov = v["coverage"]
    row = _write_title(ws, row, "Coverage / ambiguity")
    cov_df = pd.DataFrame([
        {"metric": "Total tagged", "value": cov["total_tagged"]},
        {"metric": "Relevant (signal)", "value": f"{cov['relevant']} ({cov['relevant_pct']}%)"},
        {"metric": "Filtered as noise", "value": cov["irrelevant_filtered"]},
        {"metric": "no_signal rate", "value": f"{cov['no_signal_pct']}%"},
        {"metric": "low_confidence rate", "value": f"{cov['low_confidence_pct']}%"},
        {"metric": "ambiguity rate (no_signal or low_conf)", "value": f"{cov['ambiguity_rate_pct']}%"},
    ])
    row = _write_table(ws, row, cov_df)

    row = _write_title(ws, row, "Per-theme evidence (distinct items × sources)")
    tdf = pd.DataFrame([{"theme": t["theme"], "kind": t["kind"], "strength": t["strength"],
                         "evidence_items": t["evidence_items"], "n_sources": t["n_sources"],
                         "sources": _join(t["sources"]), "triangulated": t["triangulated"]}
                        for t in v["per_theme_evidence"] if t["evidence_items"] > 0])
    row = _write_table(ws, row, tdf, wrap_cols=("sources",))

    row = _write_title(ws, row, "Per-hypothesis validation")
    hdf = pd.DataFrame([{"hypothesis": h["hypothesis"], "confidence": h["confidence"],
                         "evidence_items": h["evidence_items"], "n_sources": h["n_sources"],
                         "sources": _join(h["sources"]), "triangulated": h["source_triangulated"]}
                        for h in v["per_hypothesis_validation"]])
    if not hdf.empty:
        row = _write_table(ws, row, hdf, wrap_cols=("hypothesis", "sources"))

    row = _write_title(ws, row, "Single-source themes (flagged lower-confidence)")
    ss = v["triangulation"]["single_source_themes_flagged_low_confidence"] or ["(none)"]
    row = _write_table(ws, row, pd.DataFrame({"theme": ss}))

    mc = v["manual_check"]
    row = _write_title(ws, row, "Manual accuracy check")
    row = _write_table(ws, row, pd.DataFrame([
        {"field": "sample_size", "value": mc["sample_size"]},
        {"field": "sample_file", "value": mc["sample_file"]},
        {"field": "accuracy_rate", "value": mc["accuracy_rate"] if mc["accuracy_rate"] is not None else "(fill in after review)"},
    ]))

    _autosize(ws, {"A": 58, "B": 14, "C": 12, "D": 14, "E": 12, "F": 26, "G": 14})
    ws.freeze_panes = "A2"


def export():
    tagged = read_jsonl(TAGGED_PATH)
    if not tagged:
        raise RuntimeError("No tagged.jsonl — run the pipeline first.")
    with open(SYNTH_PATH, encoding="utf-8") as f:
        synth = json.load(f)
    with open(VALIDATION_PATH, encoding="utf-8") as f:
        validation = json.load(f)

    # Tab 1 via pandas; then Tabs 2 & 3 via openpyxl on the same workbook.
    with pd.ExcelWriter(XLSX_PATH, engine="openpyxl") as writer:
        tagged_df(tagged).to_excel(writer, sheet_name="Tagged Reviews", index=False)
        wb = writer.book
        ws1 = wb["Tagged Reviews"]
        # style tab-1 header row
        for cell in ws1[1]:
            cell.font, cell.fill = HEADER_FONT, HEADER_FILL
        ws1.freeze_panes = "A2"
        _autosize(ws1, {"A": 22, "B": 12, "C": 12, "D": 8, "E": 16, "F": 11, "G": 11,
                        "H": 20, "I": 11, "J": 26, "K": 26, "L": 20, "M": 34, "N": 34,
                        "O": 30, "P": 50, "Q": 28})
        for col in ("M", "N", "O", "P"):
            for cell in ws1[col]:
                cell.alignment = WRAP
        build_insights(wb.create_sheet("Insights"), synth)
        build_validation(wb.create_sheet("Validation"), validation)

    log.info("Wrote %s (3 tabs: Tagged Reviews, Insights, Validation)", XLSX_PATH)


if __name__ == "__main__":
    export()
